#!/usr/bin/env python3
"""
从 Excel 重新生成 red-sites-data.js，确保关联历史事件和关联人物与 Excel 一致。
同时输出图片映射信息。
"""
import openpyxl
import json
import re
import shutil
import os
import datetime

BASE = '/Users/mr.beef/工作台/数字档案馆小组作业'
EXCEL_PATH = os.path.join(BASE, '红色遗址项目/红色遗址数据.xlsx')
OUTPUT_JS = os.path.join(BASE, '网页部分/red-map/red-sites-data.js')
SITE_IMAGES_DIR = os.path.join(BASE, '网页部分/red-map/site-images')
SOURCE_IMAGES_DIR = os.path.join(BASE, '图片')

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active
headers = [str(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]

def parse_people(raw):
    """解析关联人物字段为列表 [{人物名称, 身份, 事迹简介}]"""
    if not raw or str(raw).strip() in ('待补充', '—', 'None', ''):
        return []

    raw = str(raw).strip()

    # 格式 A: R-series - 句号+空格 分隔 "Name;Role;Desc。 Name;Role;Desc。"
    if '。' in raw:
        parts = raw.split('。 ')
        parts = [p.strip().rstrip('。').strip() for p in parts if p.strip()]
        result = []
        for part in parts:
            fields = [f.strip() for f in part.split(';')]
            if not fields or not fields[0]:
                continue
            if len(fields) == 1:
                result.append({'人物名称': fields[0], '身份': '相关人物', '事迹简介': '待补充'})
            elif len(fields) == 2:
                result.append({'人物名称': fields[0], '身份': fields[1], '事迹简介': '待补充'})
            else:
                name = fields[0]
                desc = fields[-1]
                role = ';'.join(fields[1:-1])
                result.append({'人物名称': name, '身份': role, '事迹简介': desc})
        if result:
            return result

    # 格式 B: 纯分号分隔，全为短人名 "Name;Name;Name"
    fields = [f.strip() for f in raw.split(';') if f.strip()]
    all_short = all(
        len(f) <= 5 and not any(c in f for c in ['，', '。', '、', '(', '（', ' ', '　'])
        for f in fields
    )
    if all_short:
        return [{'人物名称': f, '身份': '相关人物', '事迹简介': '待补充'} for f in fields]

    # 格式 C/D: 空格分隔的人名组 "Name;Role;Desc Name;Role Name;Role;Desc"
    # 在空格+中文名(2-4字)+分号 之前分割
    parts = re.split(r'\s+(?=[一-龥]{2,4};)', raw)
    parts = [p.strip() for p in parts if p.strip()]
    if len(parts) > 1:
        result = []
        for part in parts:
            fields = [f.strip() for f in part.split(';') if f.strip()]
            if not fields:
                continue
            if len(fields) == 1:
                result.append({'人物名称': fields[0], '身份': '相关人物', '事迹简介': '待补充'})
            elif len(fields) == 2:
                result.append({'人物名称': fields[0], '身份': fields[1], '事迹简介': '待补充'})
            else:
                result.append({'人物名称': fields[0], '身份': ';'.join(fields[1:-1]), '事迹简介': fields[-1]})
        if result:
            return result

    # Fallback: 当作单人处理
    fields = [f.strip() for f in raw.split(';') if f.strip()]
    if not fields:
        return []
    if len(fields) == 1:
        return [{'人物名称': fields[0], '身份': '相关人物', '事迹简介': '待补充'}]
    elif len(fields) == 2:
        return [{'人物名称': fields[0], '身份': fields[1], '事迹简介': '待补充'}]
    else:
        return [{'人物名称': fields[0], '身份': ';'.join(fields[1:-1]), '事迹简介': fields[-1]}]


def parse_surrounding(val):
    """解析周边关联遗址"""
    if not val or str(val).strip() in ('待补充', '—', 'None', ''):
        return []
    raw = str(val).strip()
    return [x.strip() for x in raw.split(',') if x.strip()]


def format_date(val):
    """格式化时间值 — datetime 对象转为 YYYY-MM-DD，字符串保持原样"""
    if not val or str(val).strip() in ('待补充', '—', 'None', ''):
        return None
    if isinstance(val, datetime.datetime):
        return val.strftime('%Y-%m-%d')
    return str(val).strip()


features = []
image_map = {}  # R编号 -> 图片路径映射

for r in range(2, ws.max_row + 1):
    def cell(col_idx):
        return ws.cell(r, col_idx).value

    site_id = str(cell(1) or '').strip()
    site_name = str(cell(2) or '').strip()
    address = str(cell(3) or '').strip()
    lng = cell(4)
    lat = cell(5)
    site_type = str(cell(6) or '').strip()
    protection_level = str(cell(7) or '').strip()
    open_status = str(cell(8) or '').strip()
    start_time = format_date(cell(9))
    end_time = format_date(cell(10))
    event_name = str(cell(11) or '').strip()
    event_desc = str(cell(12) or '').strip()
    people_raw = cell(13)
    history_bg = str(cell(14) or '').strip()
    media_link_raw = str(cell(15) or '').strip()
    if media_link_raw in ('待补充', '—', 'None', ''):
        media_link = '待补充'
    else:
        urls = [u.strip() for u in media_link_raw.split(',') if u.strip()]
        media_link = urls if urls else '待补充'
    visit_open = str(cell(16) or '').strip()
    visit_booking = str(cell(17) or '').strip()
    visit_phone = str(cell(18) or '').strip()
    surrounding_raw = cell(19)
    data_source = str(cell(20) or '').strip()

    # 处理经纬度坐标字符串
    coord_str = ''
    if lng is not None and lat is not None:
        coord_str = f'{lng}, {lat}'
    elif lng is not None:
        coord_str = str(lng)
    elif lat is not None:
        coord_str = str(lat)

    # 解析人物
    people_list = parse_people(people_raw)

    # 解析周边遗址
    surrounding = parse_surrounding(surrounding_raw)

    # 处理事件名称为"待补充"的情况
    evt_name_final = event_name if event_name not in ('待补充', '—', 'None', '') else ''
    evt_desc_final = event_desc if event_desc not in ('待补充', '—', 'None', '') else ''

    feature = {
        "type": "Feature",
        "geometry": {
            "type": "Point",
            "coordinates": [float(lng) if lng is not None else None, float(lat) if lat is not None else None]
        },
        "properties": {
            "基础信息": {
                "遗址ID": site_id,
                "遗址名称": site_name,
                "详细地址": address,
                "经纬度坐标": coord_str,
                "遗址类型": site_type,
                "保护等级": protection_level,
                "开放状态": open_status
            },
            "史实内容": {
                "发生时间": {
                    "起始时间": start_time,
                    "结束时间": end_time
                },
                "关联历史事件": {
                    "事件名称": evt_name_final,
                    "事件概述": evt_desc_final
                },
                "关联人物": people_list,
                "历史背景与意义": history_bg if history_bg not in ('待补充', '—', 'None', '') else '',
                "影像资料与链接": media_link
            },
            "辅助信息": {
                "参观须知": {
                    "开放时间": visit_open if visit_open not in ('待补充', '—', 'None', '') else '',
                    "预约方式": visit_booking if visit_booking not in ('待补充', '—', 'None', '') else '',
                    "联系电话": visit_phone if visit_phone not in ('待补充', '—', 'None', '') else ''
                },
                "周边关联遗址": surrounding,
                "数据来源": data_source if data_source not in ('待补充', '—', 'None', '') else ''
            }
        }
    }
    features.append(feature)

    # 建立图片映射: R-001 -> 1.png, R-002 -> 2.png, ...
    if site_id.startswith('R-'):
        num_part = site_id.split('-')[1]
        try:
            num = int(num_part)
            if 1 <= num <= 96:
                src_img = os.path.join(SOURCE_IMAGES_DIR, f'{num}.png')
                if os.path.exists(src_img):
                    image_map[site_id] = num
        except ValueError:
            pass

geojson = {
    "type": "FeatureCollection",
    "features": features
}

# 写入 JS 文件
js_content = f"window.RED_SITES_DATA = {json.dumps(geojson, ensure_ascii=False, indent=2)};"
with open(OUTPUT_JS, 'w', encoding='utf-8') as f:
    f.write(js_content)

print(f"已生成 {OUTPUT_JS}，共计 {len(features)} 条遗址记录。")

# 输出图片映射信息
print(f"\n图片映射 (共 {len(image_map)} 条):")
for site_id in sorted(image_map.keys()):
    num = image_map[site_id]
    print(f"  {site_id} <- {num}.png")

# 复制图片
print("\n复制图片到 site-images/ ...")
os.makedirs(SITE_IMAGES_DIR, exist_ok=True)
copied = 0
for site_id, num in image_map.items():
    src = os.path.join(SOURCE_IMAGES_DIR, f'{num}.png')
    # 格式: site-R001.png
    dst_filename = f'site-{site_id}.png'
    dst = os.path.join(SITE_IMAGES_DIR, dst_filename)
    if os.path.exists(src):
        shutil.copy2(src, dst)
        copied += 1
    else:
        print(f"  警告: 源文件不存在 {src}")

print(f"完成! 共复制 {copied} 张图片。")

# 验证几个人物解析结果
print("\n=== 人物解析验证 ===")
test_ids = ['R-001', 'R-003', 'R-020', 'R-064', 'FJ-001', 'JX-001', 'SX-001', 'SX-013']
for f in features:
    sid = f['properties']['基础信息']['遗址ID']
    if sid in test_ids:
        people = f['properties']['史实内容']['关联人物']
        print(f"\n{sid} ({f['properties']['基础信息']['遗址名称']}):")
        for i, p in enumerate(people):
            print(f"  [{i+1}] {p['人物名称']} | {p['身份']} | {p['事迹简介'][:60]}")
